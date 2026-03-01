# -*- coding: utf-8 -*-
"""정책표 DB → 엑셀 내보내기 (import와 동일한 열 구조)."""
import io
from datetime import datetime

from openpyxl import Workbook


# import_policy_from_excel.py와 동일: 33~45열(1-based) = 인덱스 32~44
HEADER_ROW = 2  # pandas header=1 → Excel 2행
DATA_START_ROW = 3
COL_START = 33   # 1-based, 통신사
COL_END = 45     # 1-based, 합산

HEADERS = [
    "통신사", "종류", "구분", "상품명", "월요금",
    "프로모1", "프로모2", "프로모3", "프로모4",
    "경품가이드", "상품권", "현금VAT포함", "합산",
]


def build_policy_xlsx(rows):
    """
    PolicyRow 리스트를 import와 호환되는 xlsx로 만든다.
    rows: PolicyRow 쿼리 결과 (telco, kind, category, product_name, month_fee,
          promo1~4, gift_guide, voucher, cash_vat, total_fee 사용)
    Returns: bytes (xlsx 파일 내용)
    """
    wb = Workbook()
    ws = wb.active
    if ws.title == "Sheet":
        ws.title = "정책표"

    # 2행: 헤더 (33~45열만 채움, 1~32열은 비워두어 import 시 45열 구조 유지)
    for c, title in enumerate(HEADERS, start=COL_START):
        ws.cell(row=HEADER_ROW, column=c, value=title)

    for row_idx, r in enumerate(rows, start=DATA_START_ROW):
        ws.cell(row=row_idx, column=COL_START + 0, value=r.telco)
        ws.cell(row=row_idx, column=COL_START + 1, value=r.kind)
        ws.cell(row=row_idx, column=COL_START + 2, value=r.category)
        ws.cell(row=row_idx, column=COL_START + 3, value=r.product_name)
        ws.cell(row=row_idx, column=COL_START + 4, value=r.month_fee)
        ws.cell(row=row_idx, column=COL_START + 5, value=r.promo1)
        ws.cell(row=row_idx, column=COL_START + 6, value=r.promo2)
        ws.cell(row=row_idx, column=COL_START + 7, value=r.promo3)
        ws.cell(row=row_idx, column=COL_START + 8, value=r.promo4)
        ws.cell(row=row_idx, column=COL_START + 9, value=r.gift_guide)
        ws.cell(row=row_idx, column=COL_START + 10, value=r.voucher)
        ws.cell(row=row_idx, column=COL_START + 11, value=r.cash_vat)
        ws.cell(row=row_idx, column=COL_START + 12, value=r.total_fee)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
